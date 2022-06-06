#!/bin/python3

import openpyxl
import datetime

import yagmail
import csv


class HouseChores(object):
    """make next weeks timetable for roommates with existing timetable"""

    def __init__(self, chores_file, new_table):
        '''initialize the class '''
        self.chores_file = chores_file
        self.new_table = new_table
        self.read_file()

    def read_file(self):
        '''Gets the current lists of respective chores from old timetable
        spreadsheet and store them in variables
        '''
        # read workbook
        self.wb = openpyxl.load_workbook(self.chores_file)
        self.ws = self.wb.active

        self.old_mopping = []
        self.old_greens = []
        self.old_water = []

        for rowNumber in range(5, self.ws.max_row+1):  # skip first five rows
            self.mopper_name = self.ws.cell(
                row=rowNumber, column=4).value  # mopping
            self.old_mopping.append(self.mopper_name)
            self.greens_name = self.ws.cell(
                row=rowNumber, column=5).value  # greens
            self.old_greens.append(self.greens_name)
            self.fetcher_name = self.ws.cell(
                row=rowNumber, column=8).value  # water
            self.old_water.append(self.fetcher_name)
        self.__make_new_names()

    def __make_new_names(self):
        '''Uses present timetable to make list of next weeks list.
           Using private method'''

        self.new_mopping = []
        self.new_greens = []
        self.new_water = []

        # for greens
        self.new_greens += self.old_greens[-4:]
        self.new_greens += self.old_greens[-4:-1]

        # for mopping
        self.new_mopping += self.old_mopping[-4:]
        self.new_mopping += self.old_mopping[-4:-1]

        # for water
        self.new_water += self.old_water[-4:]
        self.new_water += self.old_water[-4:-1]

    def create_table(self):
        "write the new timetable to a file"

        # add duration of timetable
        self.today = datetime.date.today()
        self.duration = datetime.timedelta(days=7)
        self.end_date = self.today+self.duration
        self.today = self.today.strftime("%d-%b-%Y")
        self.end_date = self.end_date.strftime("%d-%b-%Y")
        self.ws.cell(row=2, column=1).value = 'Dated ' + \
            str(self.today)+" to "+str(self.end_date)

        self.item_number = 0
        for rowNumber in range(5, self.ws.max_row+1):  # skip first five rows
            # column 4 cleaning
            self.ws.cell(
                row=rowNumber, column=4).value = self.new_mopping[self.item_number]
            self.ws.cell(
                row=rowNumber, column=5).value = self.new_greens[self.item_number]  # greens
            self.ws.cell(
                row=rowNumber, column=8).value = self.new_water[self.item_number]  # water
            self.item_number += 1

        self.wb.save(self.new_table)
        self.wb.close()
        return self.new_table

    def notification(self, sender_email, password, new_table, phone="room.csv"):
        '''Send email notification to members
        param:sender_email,password,new timetable,phonebook csv
          '''

        self.message = """\
        Dear {name},here's this weeks HouseChores Schedule.
        """

        with open(phone) as contacts:
            self.yag = yagmail.SMTP(sender_email, password)
            self.reader = csv.reader(contacts)
            next(self.reader)  # skip head
            for name, email in self.reader:
                self.body = self.message.format(name=name)
                print(f"Sending mail to {name}")
                self.yag.send(
                    to=email,
                    subject='House chores timetable',
                    contents=[self.body, new_table]
                )
